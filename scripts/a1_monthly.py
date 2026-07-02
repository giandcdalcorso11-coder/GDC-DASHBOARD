"""
GDC IA TEAM — Script A1 Monthly
Ogni 2° del mese alle 08:00 ora italiana.

Flusso:
1. Instagram Graph API → metriche post + reel del mese precedente (originali)
2. CSV Meta Business Suite (Drive "Archivio docs MBS") → repost/menzioni del mese precedente
3. Compila Instagram_Analytics_GDC.xlsx (5 sheet)
4. Carica su Google Drive (cartella A1.2)
5. Aggiorna Supabase → stato A1 = done
6. Invia Web Push notification

Secrets GitHub richiesti:
  IG_ACCESS_TOKEN       — token a lunga durata Instagram Graph API
  IG_USER_ID            — ID numerico account Instagram
  GOOGLE_CREDENTIALS    — JSON service account Google (base64)
  DRIVE_FOLDER_A1       — ID cartella Drive A1.2
  SUPABASE_URL          — URL progetto Supabase
  SUPABASE_KEY          — anon/service key Supabase
  PUSH_SUBSCRIPTION     — JSON subscription Web Push (base64)
  VAPID_PRIVATE_KEY     — chiave privata VAPID
  VAPID_EMAIL           — email per VAPID
"""

import os
import json
import base64
import tempfile
import re
import csv
import io
import requests
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from calendar import monthrange

import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from supabase import create_client
from pywebpush import webpush, WebPushException


# ─── CONFIG ────────────────────────────────────────────────────────────────────

IG_TOKEN     = os.environ["IG_ACCESS_TOKEN"]
IG_USER_ID   = os.environ["IG_USER_ID"]
DRIVE_FOLDER = os.environ["DRIVE_FOLDER_A1"]
SUPA_URL     = os.environ["SUPABASE_URL"]
SUPA_KEY     = os.environ["SUPABASE_KEY"]
VAPID_PRIV   = os.environ["VAPID_PRIVATE_KEY"]
VAPID_EMAIL  = os.environ["VAPID_EMAIL"]

EXCEL_FILENAME       = "Instagram_Analytics_GDC"
EXCEL_FILENAME_LOCAL = "Instagram_Analytics_GDC.xlsx"

# Mese precedente (quello da analizzare)
today         = date.today()
target        = today - relativedelta(months=1)
MESE_NUM      = target.month
ANNO          = target.year
MESE_IT       = [
    "", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"
][MESE_NUM]
MESE_LABEL    = f"{MESE_IT} {ANNO}"
PRIMO_GIORNO  = date(ANNO, MESE_NUM, 1)
ULTIMO_GIORNO = date(ANNO, MESE_NUM, monthrange(ANNO, MESE_NUM)[1])

# Palette colori GDC
COLOR_GRAY_SEP = "D3D3D3"
COLOR_YELLOW   = "FFD700"
COLOR_HEADER   = "1C1C1C"
COLOR_WHITE    = "FFFFFF"

print(f"▶ A1 Monthly — {MESE_LABEL}")
print(f"  Periodo: {PRIMO_GIORNO} → {ULTIMO_GIORNO}")


# ─── 1. INSTAGRAM GRAPH API ─────────────────────────────────────────────────────

def ig_get(endpoint, params={}):
    """Chiamata GET all'API Instagram Graph."""
    base = f"https://graph.facebook.com/v21.0/{endpoint}"
    params["access_token"] = IG_TOKEN
    r = requests.get(base, params=params)
    r.raise_for_status()
    return r.json()


def fetch_profile_data():
    """
    Recupera dati aggregati del profilo: follower, following, media count.
    Richiede instagram_basic — disponibile con il token attuale.
    """
    print("  → Fetch dati profilo...")
    try:
        data = ig_get(IG_USER_ID, {
            "fields": "followers_count,follows_count,media_count,name"
        })
        followers = data.get("followers_count", "")
        following = data.get("follows_count", "")
        media_count = data.get("media_count", "")
        print(f"    Profilo: {followers} follower, {following} following, {media_count} post totali")
        return followers, following, media_count
    except Exception as e:
        print(f"    Profilo non disponibile: {e}")
        return "", "", ""


def fetch_post_insights(media_id, media_type):
    """
    Recupera metriche avanzate per un singolo post/reel via /insights.
    Solo per media propri — non disponibile per media di altri account.
    """
    # Nota (step 20, 3.4/3.12): impressions -> views. follows rimossa dalla
    # richiesta per i post: non supportata via API (confermato dai test su
    # VIDEO/Reel e CAROUSEL_ALBUM), la colonna resta strutturalmente vuota.
    if media_type in ("IMAGE", "CAROUSEL_ALBUM"):
        metric = "views,reach,saved,shares,likes,comments"
    elif media_type in ("VIDEO", "REELS"):
        metric = "views,reach,saved,shares,likes,comments,plays"
    else:
        metric = "views,reach,saved,shares,likes,comments"

    try:
        ins = ig_get(f"{media_id}/insights", {
            "metric": metric,
            "period": "lifetime"
        })
        return {m["name"]: m["values"][0]["value"] for m in ins.get("data", [])}
    except Exception:
        try:
            ins = ig_get(f"{media_id}/insights", {
                "metric": "views,reach,saved,shares",
                "period": "lifetime"
            })
            return {m["name"]: m["values"][0]["value"] for m in ins.get("data", [])}
        except Exception:
            return {}


def fetch_posts():
    """
    Recupera tutti i post/reel ORIGINALI del mese target da /me/media.
    Paginazione automatica finché non usciamo dal periodo.
    """
    print("  → Fetch post originali...")
    fields = "id,timestamp,media_type,caption,permalink,like_count,comments_count,username"
    posts = []
    url = f"{IG_USER_ID}/media"
    params = {"fields": fields, "limit": 100}

    while True:
        data = ig_get(url, params)
        items = data.get("data", [])

        for item in items:
            ts = datetime.fromisoformat(item["timestamp"].replace("Z", "+00:00"))
            item_date = ts.date()

            if item_date > ULTIMO_GIORNO:
                continue
            if item_date < PRIMO_GIORNO:
                print(f"    Trovati {len(posts)} post originali nel periodo")
                return posts

            metrics = fetch_post_insights(item["id"], item.get("media_type", "IMAGE"))
            item.update(metrics)
            posts.append(item)

        next_url = data.get("paging", {}).get("next")
        if not next_url:
            break
        import urllib.parse as urlparse
        parsed = urlparse.urlparse(next_url)
        params = dict(urlparse.parse_qsl(parsed.query))
        url = f"{IG_USER_ID}/media"

    print(f"    Trovati {len(posts)} post originali nel periodo")
    return posts


MBS_FOLDER_ID   = "1rGgK2yB_MRMi0jvxKWPSIJJJKCtgDnIg"  # Drive "Archivio docs MBS" — struttura flat
MBS_FILE_SUFFIX = "_3909528329355109.csv"              # File Content/Post — unico che importiamo

# Mappa "Tipo di post" (etichette MBS) -> valori media_type stile API,
# cosi' compile_sheet_post puo' riusare media_type_to_gdc() senza differenze
# di formattazione tra post originali e repost.
MBS_TIPO_MAP = {
    "Carosello di Instagram": "CAROUSEL_ALBUM",
    "Reel di Instagram":      "VIDEO",
    "Foto di Instagram":      "IMAGE",
    "Post di Instagram":      "IMAGE",
}


def find_mbs_csv(drive_service):
    """
    Cerca in 'Archivio docs MBS' il CSV Content/Post del mese target.
    Riconosce il file dal suffisso fisso nel nome + dalle date embedded
    (es. Jun-01-2026_Jun-30-2026_3909528329355109.csv) — non serve
    rinominare il file caricato. Se piu' file corrispondono allo stesso
    mese, usa quello caricato piu' di recente (vedi step 20, 3.7).
    """
    query = f"'{MBS_FOLDER_ID}' in parents and trashed=false and mimeType='text/csv'"
    results = drive_service.files().list(
        q=query, fields="files(id, name, createdTime)"
    ).execute()

    candidates = []
    for f in results.get("files", []):
        name = f["name"]
        if not name.endswith(MBS_FILE_SUFFIX):
            continue
        m = re.match(r"^([A-Za-z]{3})-(\d{2})-(\d{4})_", name)
        if not m:
            continue
        try:
            file_start = datetime.strptime(f"{m.group(1)} {m.group(3)}", "%b %Y").date()
        except ValueError:
            continue
        if file_start.year == ANNO and file_start.month == MESE_NUM:
            candidates.append(f)

    if not candidates:
        return None

    candidates.sort(key=lambda f: f.get("createdTime", ""), reverse=True)
    return candidates[0]["id"]


def fetch_mbs_reposts(drive_service):
    """
    Legge il CSV Content/Post MBS del mese target, tiene solo le righe
    con account diverso dal proprio (repost/menzioni — le righe del
    proprio account sono gia' coperte dall'API) e le converte nello
    stesso formato usato per i post originali (vedi step 20, 4.2).

    Se il file non e' ancora su Drive la run prosegue comunque, senza
    repost per questo giro — nessun errore bloccante.
    Notifica email + flag a1_mbs_missing su Supabase: Blocco 3, punto 7,
    ancora DA FARE — non implementato in questo step.
    """
    print("  → Fetch repost da CSV MBS...")
    file_id = find_mbs_csv(drive_service)
    if not file_id:
        print(f"    ⚠ CSV MBS non trovato per {MESE_LABEL} — repost vuoti per questo run")
        return []

    from googleapiclient.http import MediaIoBaseDownload
    request = drive_service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)
    text = fh.read().decode("utf-8-sig")

    def to_int(v):
        v = (v or "").strip()
        return int(v) if v.lstrip("-").isdigit() else 0

    reader = csv.DictReader(io.StringIO(text))
    repost = []
    for row in reader:
        account_id = (row.get("ID dell'account") or "").strip()
        if not account_id or account_id == IG_USER_ID:
            continue  # riga del proprio account — gia' coperta dall'API

        try:
            pub = datetime.strptime(row.get("Orario di pubblicazione", "").strip(), "%m/%d/%Y %H:%M")
        except ValueError:
            continue
        if not (PRIMO_GIORNO <= pub.date() <= ULTIMO_GIORNO):
            continue

        tipo_raw = row.get("Tipo di post", "").strip()
        repost.append({
            "timestamp":      pub.strftime("%Y-%m-%dT%H:%M:%S+0000"),
            "media_type":     MBS_TIPO_MAP.get(tipo_raw, tipo_raw),
            "caption":        row.get("Descrizione", ""),
            "permalink":      row.get("Permalink", ""),
            "username":       row.get("Nome utente dell'account", ""),
            "owner_name":     row.get("Nome account", ""),
            "like_count":     to_int(row.get("Mi piace")),
            "comments_count": to_int(row.get("Commenti")),
            "shares":         to_int(row.get("Condivisioni")),
            "views":          to_int(row.get("Visualizzazioni")),
            "_tipo_autore":   "REPOST",
        })

    print(f"    Trovati {len(repost)} repost nel periodo (fonte: MBS)")
    return repost


# ─── 2. GOOGLE DRIVE ────────────────────────────────────────────────────────────

def get_drive_service():
    """Crea servizio Google Drive da credenziali service account."""
    creds_b64  = os.environ["GOOGLE_CREDENTIALS"]
    creds_json = json.loads(base64.b64decode(creds_b64))
    scopes = [
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets"
    ]
    creds = Credentials.from_service_account_info(creds_json, scopes=scopes)
    return build("drive", "v3", credentials=creds), creds


def drive_find_file(service, name, folder_id):
    """Trova un file per nome in una cartella Drive. Ritorna (id, mimeType) o (None, None)."""
    query = f"name='{name}' and '{folder_id}' in parents and trashed=false"
    results = service.files().list(q=query, fields="files(id, name, mimeType)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"], files[0]["mimeType"]

    name_no_ext = name.rsplit(".", 1)[0] if "." in name else None
    if name_no_ext and name_no_ext != name:
        query2 = f"name='{name_no_ext}' and '{folder_id}' in parents and trashed=false"
        results2 = service.files().list(q=query2, fields="files(id, name, mimeType)").execute()
        files2 = results2.get("files", [])
        if files2:
            return files2[0]["id"], files2[0]["mimeType"]

    return None, None


def drive_download_excel(service, file_id, mime_type, dest_path):
    """Scarica il file Excel dal Drive. Se Foglio Google nativo, esporta in xlsx."""
    from googleapiclient.http import MediaIoBaseDownload
    import io

    GOOGLE_SHEETS_MIME = "application/vnd.google-apps.spreadsheet"
    XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if mime_type == GOOGLE_SHEETS_MIME:
        request = service.files().export_media(fileId=file_id, mimeType=XLSX_MIME)
        print(f"    Export Foglio Google → xlsx")
    else:
        request = service.files().get_media(fileId=file_id)

    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)
    with open(dest_path, "wb") as f:
        f.write(fh.read())
    print(f"    Excel scaricato: {dest_path}")


def read_stories_from_sheet(drive_service):
    """
    Legge le stories del mese target dal tab Stories_{MESE_IT}_{ANNO}
    del Google Sheet A1 (stesso Sheet in cui le scrive il daily script).
    Restituisce lista di dict compatibili con compile_sheet_stories().
    """
    from googleapiclient.discovery import build as gbuild
    from google.oauth2.service_account import Credentials as Creds

    tab_name = f"Stories_{MESE_IT}_{ANNO}"   # es. Stories_Giugno_2026
    print(f"  → Leggo stories dal Sheet tab '{tab_name}'...")

    # Riusa le credenziali già disponibili
    creds_b64  = os.environ["GOOGLE_CREDENTIALS"]
    creds_json = json.loads(base64.b64decode(creds_b64))
    scopes     = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds      = Creds.from_service_account_info(creds_json, scopes=scopes)
    sheets_svc = gbuild("sheets", "v4", credentials=creds)

    SHEET_ID = "1puwwEmieMPGIaY_xgBPO682HCZKDcIlDvJOWdu2lz30"

    try:
        result = sheets_svc.spreadsheets().values().get(
            spreadsheetId=SHEET_ID,
            range=f"'{tab_name}'!A1:M"
        ).execute()
    except Exception as e:
        print(f"    Tab '{tab_name}' non trovato o non accessibile: {e}")
        return []

    rows = result.get("values", [])
    if len(rows) < 2:
        print(f"    Tab '{tab_name}' vuoto o solo header")
        return []

    headers = rows[0]
    stories = []
    for row in rows[1:]:
        # Padda la riga se ha meno colonne dell'header
        padded = row + [""] * (len(headers) - len(row))
        s = dict(zip(headers, padded))
        # Normalizza i tipi numerici
        for field in ("views", "reach", "replies", "navigation",
                      "profile_visits", "follows", "shares"):
            try:
                s[field] = int(s[field]) if s.get(field) else 0
            except (ValueError, TypeError):
                s[field] = 0
        # Normalizza timestamp per parse_ts()
        if s.get("timestamp"):
            s["timestamp"] = s["timestamp"]
        stories.append(s)

    print(f"    Caricate {len(stories)} stories dal tab")
    return stories


def drive_upload_excel(service, local_path, folder_id, existing_id=None):
    """Carica (o aggiorna) il file Excel su Drive come Foglio Google nativo."""
    media = MediaFileUpload(
        local_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    if existing_id:
        # Specifica mimeType nel body per forzare la conversione in Google Sheets nativo
        # Senza questo Google Drive restituisce 500 quando il file è già nativo
        service.files().update(
            fileId=existing_id,
            body={"mimeType": "application/vnd.google-apps.spreadsheet"},
            media_body=media
        ).execute()
        print(f"    Foglio Google aggiornato su Drive (ID: {existing_id})")
    else:
        meta = {
            "name": EXCEL_FILENAME,
            "parents": [folder_id],
            "mimeType": "application/vnd.google-apps.spreadsheet"
        }
        f = service.files().create(body=meta, media_body=media, fields="id").execute()
        print(f"    Foglio Google creato su Drive (ID: {f['id']})")


# ─── 3. EXCEL — STRUTTURA E COMPILAZIONE ────────────────────────────────────────

SHEET_NAMES = [
    "Panoramica Profilo",
    "Insights Post",
    "Insights Stories",
    "KPI Medi",
    "Note Strategiche"
]

HEADERS = {
    "Panoramica Profilo": [
        "Mese", "Data Run", "Periodo",
        "Post Totali Profilo", "Follower Totali", "Following",
        "Visualizzazioni Totali", "Interazioni Totali",
        "N° Post/Reel Pubblicati", "N° Stories Pubblicate", "Note"
    ],
    "Insights Post": [
        "Mese", "Data", "Ora", "Tipo Contenuto", "Tipo Autore",
        "Caption", "Categoria Primaria", "Categoria Secondaria",
        "Permalink", "Visualizzazioni", "Reach", "Like", "Commenti",
        "Salvataggi", "Condivisioni", "Interazioni Totali",
        "Engagement Rate", "Tasso Viralità", "Sentiment Score",
        "Follower Acquisiti", "Collaborazione Sì/No", "Note"
    ],
    "Insights Stories": [
        "Mese", "Data", "Ora", "Permalink",
        "Visualizzazioni", "Reach", "Like", "Condivisioni",
        "Risposte", "Navigation Totale", "Sticker Taps",
        "Visite Profilo", "Follower Acquisiti", "Note"
    ],
    "KPI Medi": [
        "Mese", "ER Medio", "Viz Medie Post", "Reach Medio Post",
        "Like Medio", "Commenti Medi", "Salvataggi Medi",
        "Tasso Viralità Medio", "Sentiment Score Medio",
        "Viz Medie Stories", "Reach Medio Stories"
    ]
}


def safe_mean(values):
    """Media ignorando None e 0."""
    valid = [v for v in values if v is not None and v != ""]
    return round(sum(valid) / len(valid), 4) if valid else ""


def media_type_to_gdc(media_type):
    mapping = {"VIDEO": "Reel", "CAROUSEL_ALBUM": "Carosello", "IMAGE": "Post"}
    return mapping.get(media_type, media_type)


def parse_ts(ts_str):
    """Parsing timestamp ISO → (date_str GG/MM/AAAA, ora_str HH:MM)."""
    dt = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
    return dt.strftime("%d/%m/%Y"), dt.strftime("%H:%M")


def get_or_create_workbook(local_path):
    """Apre l'Excel esistente o ne crea uno nuovo con i 5 sheet."""
    if os.path.exists(local_path):
        wb = openpyxl.load_workbook(local_path)
        for name in SHEET_NAMES:
            if name not in wb.sheetnames:
                wb.create_sheet(name)
        print("    Workbook esistente caricato")
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for name in SHEET_NAMES:
            wb.create_sheet(name)
        for sheet_name, headers in HEADERS.items():
            ws = wb[sheet_name]
            ws.append(headers)
            _style_header_row(ws, headers)
        print("    Nuovo workbook creato")
    return wb


def _style_header_row(ws, headers):
    """Stile header: sfondo nero, testo bianco, bold."""
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    white_font  = Font(color=COLOR_WHITE, bold=True)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 18


def compile_sheet_post(wb, posts):
    """
    Compila Sheet 2 — Insights Post.
    Riceve la lista unificata di originali + repost già mergiata.
    Per i repost: impressioni e reach vuoti (non disponibili via API per media altrui).
    Per gli originali: tutte le metriche da /insights.
    Celle categoria gialle solo per ORIGINALE.
    """
    ws = wb["Insights Post"]
    headers = HEADERS["Insights Post"]

    # Controlla se il mese esiste già
    mese_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == MESE_LABEL:
            mese_rows.append(i)

    if mese_rows:
        # Controlla se ci sono repost già inseriti
        repost_presenti = any(
            ws.cell(row=r, column=5).value == "REPOST"
            for r in mese_rows
        )
        repost_in_arrivo = any(p.get("_tipo_autore") == "REPOST" for p in posts)

        if repost_presenti or not repost_in_arrivo:
            print(f"    Sheet Post: {MESE_LABEL} già presente, skip")
            return
        else:
            # Mese presente ma senza repost e ora ne abbiamo — cancella e reinserisce
            print(f"    Sheet Post: {MESE_LABEL} presente senza repost — forzo reinserimento")
            # Trova anche la riga separatore (grigia) dopo il blocco mese
            last_row = max(mese_rows)
            # Cancella dal primo al separatore incluso
            ws.delete_rows(min(mese_rows), last_row - min(mese_rows) + 2)

    posts_sorted = sorted(posts, key=lambda p: p.get("timestamp", ""), reverse=True)

    insert_at = 2
    rows_to_insert = []
    yellow_fill = PatternFill("solid", fgColor=COLOR_YELLOW)

    for p in posts_sorted:
        data_str, ora_str = parse_ts(p.get("timestamp", ""))
        tipo_contenuto = media_type_to_gdc(p.get("media_type", ""))

        # Determina tipo autore:
        # - _tipo_autore="REPOST" → da CSV MBS (fetch_mbs_reposts)
        # - username == "giandcdalcorso" → originale (fetch_posts)
        # - altrimenti → originale (fallback sicuro)
        if p.get("_tipo_autore") == "REPOST":
            tipo_autore = "REPOST"
        else:
            username    = p.get("username", p.get("owner", {}).get("username", ""))
            tipo_autore = "ORIGINALE" if (not username or username == "giandcdalcorso") else "REPOST"

        if tipo_autore == "REPOST":
            # Salvataggi, copertura e follower acquisiti non sono disponibili
            # per contenuti di altri account, nemmeno via MBS (step 20, 4.2)
            views  = p.get("views", "") or ""
            reach  = ""
            saved  = ""
            shares = p.get("shares", "") or ""
            follows = ""
        else:
            views  = p.get("views", 0) or 0
            reach  = p.get("reach") or ""
            saved  = p.get("saved", 0) or 0
            shares = p.get("shares", 0) or 0
            follows = p.get("follows", 0) or 0

        likes    = p.get("like_count", 0) or 0
        comments = p.get("comments_count", 0) or 0

        # KPI calcolati solo per originali con reach
        if tipo_autore == "ORIGINALE" and reach:
            interactions = likes + comments + (saved or 0) + (shares or 0)
            er           = round(interactions / reach, 4)
            viralita     = round(((shares or 0) / reach) * 100, 4)
        else:
            interactions = likes + comments + (saved or 0) + (shares or 0) if tipo_autore == "REPOST" else likes + comments + (saved or 0) + (shares or 0)
            er           = ""
            viralita     = ""

        sentiment = round(((saved or 0) + (shares or 0)) / likes, 4) if (tipo_autore == "ORIGINALE" and likes > 0) else ""
        collaborazione = "Sì" if tipo_autore == "REPOST" else "No"
        note = f"Repost di: {p.get('owner_name', '')}" if tipo_autore == "REPOST" and p.get("owner_name") else ""

        rows_to_insert.append([
            MESE_LABEL, data_str, ora_str, tipo_contenuto, tipo_autore,
            p.get("caption", "")[:500] if p.get("caption") else "",
            "",  # Categoria Primaria — compila tu
            "",  # Categoria Secondaria — compila tu
            p.get("permalink", ""),
            views, reach,
            likes, comments, saved, shares,
            interactions, er, viralita, sentiment,
            follows, collaborazione, note
        ])

    ws.insert_rows(insert_at, amount=len(rows_to_insert) + 1)

    for i, row_data in enumerate(rows_to_insert):
        r = insert_at + i
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=j, value=val)
            if j in (7, 8) and row_data[4] == "ORIGINALE":
                cell.fill = yellow_fill

    gray_fill = PatternFill("solid", fgColor=COLOR_GRAY_SEP)
    sep_row_num = insert_at + len(rows_to_insert)
    for col in range(1, len(headers) + 1):
        ws.cell(row=sep_row_num, column=col).fill = gray_fill

    print(f"    Sheet Post: {len(rows_to_insert)} righe inserite")
    return rows_to_insert


def compile_sheet_stories(wb, stories):
    """Compila Sheet 3 — Insights Stories (top 10 per visualizzazioni)."""
    ws = wb["Insights Stories"]
    headers = HEADERS["Insights Stories"]

    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == MESE_LABEL:
            print(f"    Sheet Stories: {MESE_LABEL} già presente, skip")
            return

    top10 = sorted(stories, key=lambda s: s.get("views", 0), reverse=True)[:10]

    insert_at = 2
    ws.insert_rows(insert_at, amount=len(top10) + 1)

    for i, s in enumerate(top10):
        data_str, ora_str = parse_ts(s.get("timestamp", ""))
        r = insert_at + i
        ws.cell(row=r, column=1,  value=MESE_LABEL)
        ws.cell(row=r, column=2,  value=data_str)
        ws.cell(row=r, column=3,  value=ora_str)
        ws.cell(row=r, column=4,  value=s.get("permalink", ""))
        ws.cell(row=r, column=5,  value=s.get("views", 0))
        ws.cell(row=r, column=6,  value=s.get("reach", ""))
        ws.cell(row=r, column=7,  value=s.get("like_count", 0))
        ws.cell(row=r, column=8,  value=s.get("shares", 0))
        ws.cell(row=r, column=9,  value=s.get("replies", 0))
        ws.cell(row=r, column=10, value=s.get("navigation", 0))
        ws.cell(row=r, column=11, value="")  # Sticker Taps: non piu' scomponibile via API (navigation e' aggregato unico) — colonna da valutare in pulizia (3.11)
        ws.cell(row=r, column=12, value=s.get("profile_visits", 0))
        ws.cell(row=r, column=13, value=s.get("follows", 0))
        ws.cell(row=r, column=14, value="")

    gray_fill = PatternFill("solid", fgColor=COLOR_GRAY_SEP)
    sep_row = insert_at + len(top10)
    for col in range(1, len(headers) + 1):
        ws.cell(row=sep_row, column=col).fill = gray_fill

    print(f"    Sheet Stories: {len(top10)} righe inserite")
    return top10


def compile_sheet_panoramica(wb, posts, stories, profile_data=None):
    """Compila Sheet 1 — Panoramica Profilo."""
    ws = wb["Panoramica Profilo"]

    # Cancella e reinserisce sempre — dati profilo e totali possono cambiare
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
        if row[0] == MESE_LABEL:
            ws.delete_rows(i, 1)
            break

    originali = [p for p in posts if p.get("_tipo_autore") != "REPOST" and
                 p.get("username", "giandcdalcorso") == "giandcdalcorso"]
    repost    = [p for p in posts if p.get("_tipo_autore") == "REPOST"]

    tot_views_post    = sum(p.get("views", 0) or 0 for p in originali)
    tot_views_stories = sum(s.get("views", 0) or 0 for s in stories)
    tot_interactions  = sum(
        (p.get("like_count", 0) or 0) +
        (p.get("comments_count", 0) or 0) +
        (p.get("saved", 0) or 0) +
        (p.get("shares", 0) or 0)
        for p in posts
    )
    periodo = f"{PRIMO_GIORNO.strftime('%d/%m/%Y')} → {ULTIMO_GIORNO.strftime('%d/%m/%Y')}"

    followers, following, media_count = profile_data if profile_data else ("", "", "")

    row_data = [
        MESE_LABEL,
        date.today().strftime("%d/%m/%Y"),
        periodo,
        media_count,
        followers,
        following,
        tot_views_post + tot_views_stories,
        tot_interactions,
        len(originali),
        len(stories),
        f"Repost taggati: {len(repost)}"
    ]

    ws.insert_rows(2, amount=1)
    for j, val in enumerate(row_data, 1):
        ws.cell(row=2, column=j, value=val)

    print(f"    Sheet Panoramica: riga {MESE_LABEL} inserita (originali: {len(originali)}, repost: {len(repost)})")


def compile_sheet_kpi(wb, posts, stories):
    """
    Compila Sheet 4 — KPI Medi.
    I KPI sono calcolati solo sui post ORIGINALI con reach disponibile.
    I repost non entrano nei calcoli (reach/impressioni non disponibili via API).
    """
    ws = wb["KPI Medi"]

    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == MESE_LABEL:
            print(f"    Sheet KPI: {MESE_LABEL} già presente, skip")
            return

    originali = [p for p in posts if p.get("_tipo_autore") != "REPOST" and
                 p.get("username", "giandcdalcorso") == "giandcdalcorso"]
    con_reach  = [p for p in originali if p.get("reach")]

    er_vals       = []
    viralita_vals = []
    for p in con_reach:
        r = p["reach"]
        i = (p.get("like_count", 0) or 0) + (p.get("comments_count", 0) or 0) + \
            (p.get("saved", 0) or 0) + (p.get("shares", 0) or 0)
        s = p.get("shares", 0) or 0
        er_vals.append(round(i / r, 4) if r else None)
        viralita_vals.append(round((s / r) * 100, 4) if r else None)

    sentiment_vals = []
    for p in originali:
        likes  = p.get("like_count", 0) or 0
        saved  = p.get("saved", 0) or 0
        shares = p.get("shares", 0) or 0
        if likes > 0:
            sentiment_vals.append(round((saved + shares) / likes, 4))

    row_data = [
        MESE_LABEL,
        safe_mean(er_vals),
        safe_mean([p.get("views", 0) for p in originali]),
        safe_mean([p.get("reach") for p in con_reach]),
        safe_mean([p.get("like_count", 0) for p in originali]),
        safe_mean([p.get("comments_count", 0) for p in originali]),
        safe_mean([p.get("saved", 0) for p in originali]),
        safe_mean(viralita_vals),
        safe_mean(sentiment_vals),
        safe_mean([s.get("views", 0) for s in stories]),
        safe_mean([s.get("reach", 0) for s in stories]),
    ]

    ws.insert_rows(2, amount=1)
    for j, val in enumerate(row_data, 1):
        ws.cell(row=2, column=j, value=val)

    print(f"    Sheet KPI: riga {MESE_LABEL} inserita")


# ─── 4. SUPABASE ─────────────────────────────────────────────────────────────────

def supabase_update_state(agent_id, stato, extra={}):
    """Aggiorna lo stato di un agente nella tabella agent_states."""
    supabase = create_client(SUPA_URL, SUPA_KEY)
    data = {
        "agent_id": agent_id,
        "stato": stato,
        "updated_at": datetime.utcnow().isoformat(),
        **extra
    }
    supabase.table("agent_states").upsert(data, on_conflict="agent_id").execute()
    print(f"    Supabase: {agent_id} → {stato}")


# ─── 5. WEB PUSH ─────────────────────────────────────────────────────────────────

def send_push(title, body):
    """Invia Web Push notification via VAPID."""
    sub_b64 = os.environ.get("PUSH_SUBSCRIPTION", "")
    if not sub_b64:
        print("    Push: nessuna subscription configurata, skip")
        return
    try:
        subscription = json.loads(base64.b64decode(sub_b64))
        webpush(
            subscription_info=subscription,
            data=json.dumps({"title": title, "body": body}),
            vapid_private_key=VAPID_PRIV,
            vapid_claims={"sub": f"mailto:{VAPID_EMAIL}"}
        )
        print(f"    Push inviata: {title}")
    except WebPushException as e:
        print(f"    Push ERRORE: {e}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────────

def main():
    supabase_update_state("a1", "working")

    # 1. Fetch dati aggregati profilo (follower, following, media count)
    profile_data = fetch_profile_data()

    # 2. Fetch post originali dal profilo
    posts_originali = fetch_posts()

    # 3. Setup Drive (serve anche per leggere il CSV MBS dei repost)
    drive_service, _ = get_drive_service()

    # 4. Fetch repost da CSV MBS (sostituisce /tags — vedi step 20, 3.6)
    posts_repost = fetch_mbs_reposts(drive_service)

    # 5. Unifica: originali + repost, ordinati per timestamp
    all_posts = posts_originali + posts_repost

    # 6. Legge stories dal tab Google Sheet (salvate dallo script daily)
    stories = read_stories_from_sheet(drive_service)

    # 7. Scarica Excel esistente (o parti da zero)
    with tempfile.TemporaryDirectory() as tmpdir:
        local_excel = os.path.join(tmpdir, EXCEL_FILENAME_LOCAL)

        existing_id, existing_mime = drive_find_file(drive_service, EXCEL_FILENAME_LOCAL, DRIVE_FOLDER)
        if existing_id:
            drive_download_excel(drive_service, existing_id, existing_mime, local_excel)

        wb = get_or_create_workbook(local_excel)

        # 8. Compila i 4 sheet
        compile_sheet_post(wb, all_posts)
        compile_sheet_stories(wb, stories)
        compile_sheet_panoramica(wb, all_posts, stories, profile_data)
        compile_sheet_kpi(wb, all_posts, stories)

        # 9. Salva e carica su Drive
        wb.save(local_excel)
        drive_upload_excel(drive_service, local_excel, DRIVE_FOLDER, existing_id)

    # 10. Aggiorna Supabase → done
    supabase_update_state("a1", "done", {
        "mese": MESE_LABEL,
        "n_post": len(posts_originali),
        "n_stories": len(stories),
        "drive_folder": DRIVE_FOLDER
    })

    # 11. Push notification → sblocca A2
    supabase_update_state("a2", "ready")
    send_push(
        "✅ A1 completato",
        f"Analytics {MESE_LABEL} pronti. Apri A2 per generare il report."
    )

    print(f"\n✅ A1 Monthly completato — {MESE_LABEL}")
    print(f"   Post originali: {len(posts_originali)}")
    print(f"   Post taggati (repost): {len(posts_repost)}")
    print(f"   Stories in archivio: {len(stories)}")


if __name__ == "__main__":
    main()
